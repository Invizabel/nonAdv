#!/usr/bin/env python3
"""
Extract the most harmful CEM rollouts into an Excel sheet.

Each row in the Excel file has:
1. harm score
2. 1st user utterance
3. 1st chatbot response
4. 2nd user utterance
5. 2nd chatbot response
6. 3rd user utterance
7. 3rd chatbot response
8. user trait profile
9. user story

Examples:
python extract_harmful_rollouts_to_excel.py --rollouts rollouts.json --top-n 20

python extract_harmful_rollouts_to_excel.py --rollouts rollouts.json --threshold 0.2

python extract_harmful_rollouts_to_excel.py \
  --rollouts rollouts.json \
  --threshold 0.1 \
  --output harmful_rollouts_over_0p1.xlsx
"""

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def load_records(path):
    """
    Supports:
    - JSON list: [{...}, {...}]
    - JSONL: one JSON object per line
    """
    path = Path(path)
    text = path.read_text()

    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for key in ["records", "data", "rollouts"]:
                if key in data and isinstance(data[key], list):
                    return data[key]
            return [data]
    except json.JSONDecodeError:
        pass

    records = []
    for i, line in enumerate(text.splitlines(), start=1):
        line = line.strip()
        if not line:
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError as e:
            raise ValueError(f"Could not parse line {i} in {path}: {e}") from e

    return records


def infer_score_key(record):
    for key in ["score", "harm_score", "harmScore", "max_score"]:
        if key in record:
            return key
    raise KeyError(
        "Could not find harm score key. Expected one of: "
        "'score', 'harm_score', 'harmScore', or 'max_score'."
    )


def get_user_assistant_turns(record, max_turns=3):
    """
    Extracts up to 3 user utterances and 3 assistant responses from:
    record["sim_out"]["conversation"]

    Ignores system messages.
    """
    conversation = record.get("sim_out", {}).get("conversation", [])

    users = []
    assistants = []

    for msg in conversation:
        role = msg.get("role")
        content = msg.get("content", "")

        if role == "user":
            users.append(content)
        elif role == "assistant":
            assistants.append(content)

    # Pad so every row has exactly 3 user turns and 3 assistant turns.
    while len(users) < max_turns:
        users.append("")
    while len(assistants) < max_turns:
        assistants.append("")

    return users[:max_turns], assistants[:max_turns]


def get_profile(record):
    profile = record.get("sim_out", {}).get("profile", None)

    if profile is None:
        # Fallback if profile is not in sim_out.
        profile = record.get("traits", [])

    if isinstance(profile, list):
        return "; ".join(str(x) for x in profile)

    return str(profile)


def get_story(record):
    return record.get("sim_out", {}).get("story", "")


def select_rollouts(records, score_key, top_n=None, threshold=None):
    for rec in records:
        rec["_parsed_score"] = float(rec[score_key])

    records = sorted(records, key=lambda r: r["_parsed_score"], reverse=True)

    if top_n is not None:
        return records[:top_n]

    if threshold is not None:
        return [r for r in records if r["_parsed_score"] >= threshold]

    raise ValueError("You must provide either --top-n or --threshold.")


def write_excel(selected_records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Harmful Rollouts"

    headers = [
        "Harm Score",
        "1st User Utterance",
        "1st Chatbot Response",
        "2nd User Utterance",
        "2nd Chatbot Response",
        "3rd User Utterance",
        "3rd Chatbot Response",
        "User Trait Profile",
        "User Story",
    ]

    ws.append(headers)

    for rec in selected_records:
        users, assistants = get_user_assistant_turns(rec)

        row = [
            rec["_parsed_score"],
            users[0],
            assistants[0],
            users[1],
            assistants[1],
            users[2],
            assistants[2],
            get_profile(rec),
            get_story(rec),
        ]

        ws.append(row)

    # Styling
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_gray = Side(style="thin", color="CCCCCC")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # Number format for harm score
    for cell in ws["A"][1:]:
        cell.number_format = "0.000000"

    # Column widths
    widths = {
        1: 14,   # harm score
        2: 45,   # user 1
        3: 55,   # assistant 1
        4: 45,   # user 2
        5: 55,   # assistant 2
        6: 45,   # user 3
        7: 55,   # assistant 3
        8: 45,   # profile
        9: 70,   # story
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Make rows tall enough to read but not absurdly huge.
    ws.row_dimensions[1].height = 35
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 160

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add autofilter
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--rollouts",
        required=True,
        help="Path to rollout JSON or JSONL file.",
    )

    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--top-n",
        type=int,
        help="Number of most harmful rollouts to extract.",
    )
    group.add_argument(
        "--threshold",
        type=float,
        help="Extract all rollouts with harm score >= this threshold.",
    )

    parser.add_argument(
        "--output",
        default=None,
        help="Output Excel filepath. Defaults based on top-n or threshold.",
    )

    args = parser.parse_args()

    records = load_records(args.rollouts)

    if not records:
        raise ValueError("No rollout records found.")

    score_key = infer_score_key(records[0])

    selected = select_rollouts(
        records,
        score_key=score_key,
        top_n=args.top_n,
        threshold=args.threshold,
    )

    if args.output is not None:
        output_path = args.output
    elif args.top_n is not None:
        output_path = f"top_{args.top_n}_harmful_rollouts.xlsx"
    else:
        threshold_str = str(args.threshold).replace(".", "p")
        output_path = f"harmful_rollouts_score_gte_{threshold_str}.xlsx"

    write_excel(selected, output_path)

    print(f"Loaded {len(records)} rollouts.")
    print(f"Selected {len(selected)} rollouts.")
    print(f"Saved Excel file to: {output_path}")


if __name__ == "__main__":
    main()